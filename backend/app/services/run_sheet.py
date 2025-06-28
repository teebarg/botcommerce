import csv
import time
from datetime import datetime
from io import BytesIO

from fastapi import (
    HTTPException,
)
from openpyxl import Workbook, load_workbook

from app.services.websocket import manager
from app.core.logging import logger
from app.core.utils import generate_data_export_email, send_email

from typing import List
from datetime import datetime
from app.prisma_client import prisma as db
from app.core.deps import supabase


async def broadcast_channel(data, user_id: int):
    await manager.send_to_user(
        user_id=user_id,
        data=data,
        message_type="sheet-processor",
    )


# Export products
async def generate_excel_file(email: str) -> str:
    logger.debug("Products export started.......")

    products = await db.product.find_many(
        include={
            "categories": True,
            "collections": True,
            "brand": True,
            "images": True,
            "variants": True,
        }
    )

    if not products:
        raise HTTPException(status_code=404, detail="No products found")

    # Create a workbook and select the active worksheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"

    # Define the header row
    headers = [
        "id",
        "name",
        "slug",
        "description",
        "price",
        "old_price",
        "inventory",
        "ratings",
        "image",
        "is_active",
        "categories",
        "collections",
        "brand",
        "images",
        "variants",
    ]
    sheet.append(headers)

    # Fetch product data and append each product as a row
    for product in products:
        # Create a comma-separated string of collection names
        categories_str = ",".join(
            [category.slug for category in product.categories])
        collections_str = ",".join(
            [collection.slug for collection in product.collections])
        images_str = "|".join([image.image for image in product.images])

        # Append the product data as a row
        sheet.append(
            [
                product.id,
                product.name,
                product.slug,
                product.description,
                product.variants[0].price if product.variants else 0,
                product.variants[0].old_price if product.variants else 0,
                product.variants[0].inventory if product.variants else 0,
                product.ratings,
                product.image,
                True,
                categories_str,
                collections_str,
                product.brand.name if product.brand else "",
                images_str,
                "|".join([variant.sku for variant in product.variants]),
            ]
        )

    # Create an in-memory Excel file using BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Generate a unique filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"product_export_{timestamp}.xlsx"

    # Convert BytesIO to bytes to match the expected input type
    output_bytes = output.getvalue()
    result = supabase.storage.from_("exports").upload(filename, output_bytes, {
        "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"})

    # Get public URL
    public_url = supabase.storage.from_("exports").get_public_url(filename)

    # Send email with download link
    email_data = await generate_data_export_email(download_link=public_url)
    send_email(
        email_to=email,
        subject="Product Export Ready",
        html_content=email_data.html_content,
    )
    logger.debug("Product export complete")

    return public_url


async def parse_categories(categories_str: str) -> List[str]:
    """Parse comma-separated categories string into a list."""
    if not categories_str:
        return []
    return [cat.strip() for cat in categories_str.split(",") if cat.strip()]


async def parse_collections(collections_str: str) -> List[str]:
    """Parse comma-separated collections string into a list."""
    if not collections_str:
        return []
    return [col.strip() for col in collections_str.split(",") if col.strip()]


async def parse_images(images_str: str) -> List[str]:
    """Parse pipe-separated images string into a list."""
    if not images_str:
        return []
    return [img.strip() for img in images_str.split("|") if img.strip()]


async def upsert_brand(brand_name: str) -> int:
    """Upsert a brand and return its ID."""
    if not brand_name or not brand_name.strip():
        return None

    slug = brand_name.lower().replace(" ", "-")
    brand = await db.brand.upsert(
        where={"slug": slug},
        data={
            "create": {
                "name": brand_name,
                "slug": slug,
                "is_active": True
            },
            "update": {}
        }
    )
    return brand.id


async def upsert_category(category_name: str) -> int:
    """Upsert a category and return its ID."""
    slug = category_name.lower().replace(" ", "-")
    category = await db.category.upsert(
        where={"slug": slug},
        data={
            "create": {
                "name": category_name,
                "slug": slug,
                "is_active": True
            },
            "update": {}
        }
    )
    return category.id


async def upsert_collection(collection_name: str) -> int:
    """Upsert a collection and return its ID."""
    slug = collection_name.lower().replace(" ", "-")
    collection = await db.collection.upsert(
        where={"slug": slug},
        data={
            "create": {
                "name": collection_name,
                "slug": slug,
                "is_active": True
            },
            "update": {}
        }
    )
    return collection.id


async def bulk_upload_products(products: list[dict]):
    try:
        # Process each product
        for product_data in products:
            # Handle categories
            category_names = await parse_categories(product_data.get("categories", ""))
            category_ids = []
            for cat_name in category_names:
                cat_id = await upsert_category(cat_name)
                category_ids.append(cat_id)

            # Handle collections
            collection_names = await parse_collections(product_data.get("collections", ""))
            collection_ids = []
            for coll_name in collection_names:
                coll_id = await upsert_collection(coll_name)
                collection_ids.append(coll_id)

            # Handle brand
            brand_name = product_data.get("brand", "")
            brand_id = await upsert_brand(brand_name)

            # Handle additional images
            image_urls = await parse_images(product_data.get("images", ""))

            # Determine product status based on inventory
            # status = "IN_STOCK" if product_data["inventory"] > 0 else "OUT_OF_STOCK"
            status = "IN_STOCK"

            create_data = {
                "name": product_data["name"],
                "slug": product_data["slug"],
                "sku": product_data["sku"],
                "description": product_data["description"],
                "image": product_data["image"],
                "status": status,
                "ratings": float(product_data["ratings"]) if product_data["ratings"] else 0.0,
                "categories": {"connect": [{"id": cid} for cid in category_ids]},
                "collections": {"connect": [{"id": cid} for cid in collection_ids]},
            }

            if brand_id:
                create_data["brand"] = {"connect": {"id": brand_id}}

            # Upsert product
            product = await db.product.upsert(
                where={"slug": product_data["slug"]},
                data={
                    "create": create_data,
                    "update": {
                        "name": product_data["name"],
                        "description": product_data["description"],
                        "image": product_data["image"],
                        "status": status,
                        "brand": {"connect": {"id": brand_id}} if brand_id else {"disconnect": True},
                        "ratings": float(product_data["ratings"]) if product_data["ratings"] else 0.0,
                        "categories": {"set": [{"id": cid} for cid in category_ids]},
                        "collections": {"set": [{"id": cid} for cid in collection_ids]},
                    }
                }
            )

            # Handle product variants (assuming one variant per product for this data)
            await db.productvariant.upsert(
                where={"sku": f"{product_data['slug']}-default"},
                data={
                    "create": {
                        "product": {"connect": {"id": product.id}},
                        "sku": f"{product_data['slug']}-default",
                        "status": status,
                        "price": float(product_data["price"]),
                        "old_price": float(product_data["old_price"]) if product_data["old_price"] else 0.0,
                        "inventory": product_data["inventory"]
                    },
                    "update": {
                        "price": float(product_data["price"]),
                        "old_price": float(product_data["old_price"]) if product_data["old_price"] else 0.0,
                        "inventory": product_data["inventory"],
                        "status": status
                    }
                }
            )

            # Handle additional images
            if image_urls:
                # Clear existing images
                await db.productimage.delete_many(where={"product_id": product.id})
                # Add new images
                await db.productimage.create_many(
                    data=[
                        {"product_id": product.id, "image": url, "order": index}
                        for index, url in enumerate(image_urls)
                    ]
                )

            print(f"Processed product: {product_data['name']}")
        print("Bulk upload completed successfully")

    except Exception as e:
        print(f"Error during bulk upload: {str(e)}")


async def process_products(file_content, content_type: str, user_id: int) -> list[dict]:
    """Load product data from a TSV file."""
    start_time = time.time()  # Start timing the process
    try:
        # Create a BytesIO stream from the file content
        file_stream = BytesIO(file_content)

        if content_type in [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "text/xlsx",
        ]:
            # Read Excel file using openpyxl
            workbook = load_workbook(file_stream)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            headers = rows[0]
            data_rows = rows[1:]

        elif content_type == "text/csv":
            # Read CSV file using csv module
            file_stream.seek(0)
            csv_reader = csv.reader(file_stream.decode("utf-8").splitlines())
            rows = list(csv_reader)
            headers = rows[0]
            data_rows = rows[1:]

        else:
            raise ValueError(
                "Unsupported file format. Please upload a CSV or Excel file."
            )

        # Batch size and iteration setup
        batch_size = 20
        num_batches = (len(data_rows) // batch_size) + 1

        # Send WebSocket update
        await broadcast_channel(
            data={
                "total_rows": num_batches * batch_size,
                "processed_rows": 1,
                "status": "processing",
            },
            user_id=user_id,
        )

        for i in range(num_batches):
            batch = data_rows[i * batch_size: (i + 1) * batch_size]
            logger.info(f"Processing batch {i + 1}")

            # Extract and process the products
            products = []
            for row in batch:
                row_data = dict(zip(headers, row, strict=False))
                name = row_data.get("name", "")
                active = row_data.get("is_active", True)
                is_active = row_data.get("is_active", True)

                if isinstance(is_active, str):
                    is_active = True if active == "TRUE" else False

                product_data = {
                    "id": row_data.get("id", ""),
                    "name": name,
                    "slug": row_data.get("slug", name.lower().replace(" ", "-")),
                    "sku": row_data.get("sku", f"{name.lower().replace(' ', '-')}-default"),
                    "description": row_data.get("description", ""),
                    "price": float(row_data.get("price", 0.0)),
                    "old_price": float(row_data.get("old_price", 0.0)),
                    "inventory": row_data.get("inventory", 1),
                    "is_active": is_active,
                    "ratings": row_data.get("ratings", 4.5),
                    "image": row_data.get("image", ""),
                    "categories": row_data.get("categories", ""),
                    "collections": row_data.get("collections", ""),
                    "brand": row_data.get("brand", ""),
                    "images": row_data.get("images", ""),
                }

                products.append(product_data)

            await bulk_upload_products(products)
            print(f"Bulk upload for batch {i + 1} completed successfully")

            # Send WebSocket update
            await broadcast_channel(
                data={
                    "total_rows": num_batches * batch_size,
                    "processed_rows": (i + 1) * batch_size,
                    "status": "processing",
                },
                user_id=user_id,
            )

        logger.info("Sheet processed successfully")
        end_time = time.time()
        logger.info(
            f"Total processing time: {end_time - start_time:.2f} seconds"
        )

        return len(products)
    except Exception as e:
        logger.error(f"An error occurred while processing. Error{e}")
        await manager.send_to_user(
            user_id=str(user_id),
            data={
                "message": f"An error occurred while processing. Error{e}",
                "status": "error",
            },
            message_type="sheet-processor",
        )
