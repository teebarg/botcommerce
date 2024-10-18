import csv
from datetime import datetime
from io import BytesIO
from typing import Any, List

from fastapi import (
    HTTPException,
)
from openpyxl import Workbook, load_workbook
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.sql import delete
from sqlmodel import Session, select

from api.websocket import manager
from core.logging import logger
from core.utils import generate_data_export_email, send_email
from db.engine import engine
from models.generic import (
    Category,
    Collection,
    Product,
    ProductCategory,
    ProductCollection,
    ProductImages,
)


async def process_products(file_content, content_type: str, user_id: int):
    try:
        # Create a BytesIO stream from the file content
        file_stream = BytesIO(file_content)

        if content_type in [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "text/xlsx",
        ]:
            # Read Excel file using openpyxl
            workbook = load_workbook(file_stream)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            headers = rows[0]
            data_rows = rows[1:]

        elif content_type == "text/csv":
            # Read CSV file using csv module
            file_stream.seek(0)
            csv_reader = csv.reader(file_stream.decode("utf-8").splitlines())
            rows = list(csv_reader)
            headers = rows[0]
            data_rows = rows[1:]

        else:
            raise ValueError(
                "Unsupported file format. Please upload a CSV or Excel file."
            )

        # Batch size and iteration setup
        batch_size = 5
        # batch_size = 500
        num_batches = (len(data_rows) // batch_size) + 1
        # Track all product slugs from the sheet
        product_slugs_in_sheet = set()

        # Send WebSocket update
        await broadcast_channel(
            data={
                "total_rows": num_batches * batch_size,
                "processed_rows": 1,
                "status": "processing",
            },
            user_id=user_id,
        )

        for i in range(num_batches):
            batch = data_rows[i * batch_size : (i + 1) * batch_size]
            logger.info(f"Processing batch {i + 1}")

            # Extract and process the products
            products_to_create_or_update = []
            for row in batch:
                row_data = dict(zip(headers, row))
                name = row_data.get("name", "")

                active = row_data.get("is_active", True)
                is_active = row_data.get("is_active", True)
                if isinstance(is_active, str):
                    is_active = True if active == "TRUE" else False

                product_data = {
                    "id": row_data.get("id", ""),
                    "name": name,
                    "slug": row_data.get("slug", name.lower().replace(" ", "-")),
                    "description": row_data.get("description", ""),
                    "price": int(row_data.get("price", 0)),
                    "old_price": int(row_data.get("old_price", 0)),
                    "inventory": row_data.get("inventory", 1),
                    "is_active": is_active,
                    "ratings": row_data.get("ratings", 4.7),
                    "image": row_data.get("image", ""),
                }

                categories = row_data.get("categories", "")
                collections = row_data.get("collections", "")
                images = row_data.get("images", "")
                product_slugs_in_sheet.add(product_data["slug"])

                # Add product data to be processed later
                products_to_create_or_update.append(
                    (product_data, categories, collections, images)
                )

            # Process the batch
            await create_or_update_products_in_db(products_to_create_or_update)

            # Send WebSocket update
            await broadcast_channel(
                data={
                    "total_rows": num_batches * batch_size,
                    "processed_rows": (i + 1) * batch_size,
                    "status": "processing",
                },
                user_id=user_id,
            )

        # After processing, delete products not in the sheet
        await delete_products_not_in_sheet(
            product_slugs_in_sheet=product_slugs_in_sheet, user_id=user_id
        )

        logger.info("Sheet processed successfully")
    except Exception as e:
        logger.error(f"An error occurred while processing. Error{e}")
        await manager.broadcast(
            id="sheet",
            data={
                "message": f"An error occurred while processing. Error{e}",
                "status": "error",
            },
            type="sheet-processor",
        )


async def create_or_update_products_in_db(products: List):
    with Session(engine) as session:
        for product_data, categories, collections, images in products:
            logger.info(f"Processing product {product_data['name']}")
            try:
                # Check if the product exists
                product = session.exec(
                    select(Product).where(Product.id == product_data["id"])
                ).first()

                if product:
                    # Update existing product
                    for key, value in product_data.items():
                        setattr(product, key, value)
                    session.add(product)
                else:
                    del product_data["id"]
                    # Create new product
                    product = Product(**product_data)
                    session.add(product)

                session.commit()
                session.refresh(product)
                await update_categories(
                    product=product, categories=categories, session=session
                )
                await update_collections(product, collections, session)
                await update_images(product, images, session)

            except SQLAlchemyError as e:
                logger.error("Error updating product" + str(e))
                await manager.broadcast(
                    id="sheet",
                    data={
                        "message": f"Error processing product: {str(e)}",
                        "status": "error",
                    },
                    type="sheet-processor",
                )
                session.rollback()


async def update_categories(product, categories, session: Session):
    session.exec(
        delete(ProductCategory).where(ProductCategory.product_id == product.id)
    )
    session.commit()

    if not categories or categories is None or not isinstance(categories, str):
        return

    for item in categories.split(","):
        category = session.exec(
            select(Category).where(Category.slug == item.strip())
        ).first()
        session.add(ProductCategory(product_id=product.id, category_id=category.id))

    session.commit()


async def update_collections(product, collections, session: Session):
    session.exec(
        delete(ProductCollection).where(ProductCollection.product_id == product.id)
    )
    session.commit()

    if not collections or collections is None or not isinstance(collections, str):
        return

    for item in collections.split(","):
        collection = session.exec(
            select(Collection).where(Collection.slug == item.strip())
        ).first()
        session.add(
            ProductCollection(product_id=product.id, collection_id=collection.id)
        )

    session.commit()


async def update_images(product, images, session: Session):
    session.exec(delete(ProductImages).where(ProductImages.product_id == product.id))
    session.commit()

    if not images or images is None or not isinstance(images, str):
        return

    for item in images.split("|"):
        session.add(ProductImages(product_id=product.id, image=item.strip()))

    session.commit()


async def delete_products_not_in_sheet(product_slugs_in_sheet: set, user_id: int):
    try:
        with Session(engine) as session:
            # Query the database after all insertions and updates
            existing_slugs_in_db = set(session.exec(select(Product.slug)).all())

            # Remove products from the database that are not in the sheet
            slugs_to_remove = existing_slugs_in_db - product_slugs_in_sheet
            if slugs_to_remove:
                session.exec(delete(Product).where(Product.slug.in_(slugs_to_remove)))
                session.commit()
                logger.info(
                    f"Successfully deleted {len(product_slugs_in_sheet)} products."
                )
            else:
                logger.info("No products to delete. All products are in the sheet.")

            await broadcast_channel(
                data={
                    "message": "Products successfully synced with the sheet",
                    "status": "completed",
                },
                user_id=user_id,
            )

    except SQLAlchemyError as e:
        logger.error(f"An error occurred while deleting products. Error {e}")
        await manager.broadcast(
            id="sheet",
            data={
                "message": f"An error occurred while deleting products. Error {e}",
                "status": "error",
            },
            type="sheet-processor",
        )


# Export products
async def generate_excel_file(bucket: Any, email: str):
    logger.debug("Products export started.......")
    with Session(engine) as session:
        products = session.exec(select(Product)).all()

        if not products:
            raise HTTPException(status_code=404, detail="No products found")

        # Create a workbook and select the active worksheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Products"

        # Define the header row
        headers = [
            "id",
            "name",
            "slug",
            "description",
            "price",
            "old_price",
            "inventory",
            "ratings",
            "image",
            "is_active",
            "categories",
            "collections",
            "images",
        ]
        sheet.append(headers)

        # Fetch product data and append each product as a row
        for product in products:
            categories = session.exec(
                select(Category.slug)
                .join(ProductCategory, Category.id == ProductCategory.category_id)
                .where(ProductCategory.product_id == product.id)
            ).all()

            # Fetch product collections as a comma-separated string
            collections = session.exec(
                select(Collection.slug)
                .join(
                    ProductCollection, Collection.id == ProductCollection.collection_id
                )
                .where(ProductCollection.product_id == product.id)
            ).all()

            # Fetch product images as a |-separated string
            images = session.exec(
                select(ProductImages.image).where(
                    ProductImages.product_id == product.id
                )
            ).all()

            # Create a comma-separated string of collection names
            categories_str = ",".join(categories)
            collections_str = ",".join(collections)
            images_str = "|".join(images)

            # Append the product data as a row
            sheet.append(
                [
                    product.id,
                    product.name,
                    product.slug,
                    product.description,
                    product.price,
                    product.old_price,
                    product.inventory,
                    product.ratings,
                    product.image,
                    product.is_active,
                    categories_str,
                    collections_str,
                    images_str,
                ]
            )

        # Create an in-memory Excel file using BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        # Generate a unique filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"product_export_{timestamp}.xlsx"

        # Upload the in-memory file to Firebase
        blob = bucket.blob(f"exports/{filename}")
        blob.upload_from_file(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        blob.make_public()  # Make the file publicly accessible
        download_url = blob.public_url

        # Send email with download link
        email_data = generate_data_export_email(download_link=download_url)
        send_email(
            email_to=email,
            subject="Product Export Ready",
            html_content=email_data.html_content,
        )
        logger.debug("Product export complete")

        return download_url


async def broadcast_channel(data, user_id: int):
    await manager.broadcast(
        id=str(user_id),
        data=data,
        type="sheet-processor",
    )
